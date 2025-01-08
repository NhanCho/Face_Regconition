import cv2
import face_recognition
from flask import Flask, render_template, Response, request, jsonify, stream_with_context
from openpyxl import Workbook, load_workbook
from datetime import datetime
import numpy as np
import os
import math
import time
import json

app = Flask(__name__)

# Load known face encodings and names
known_face_encodings = []
known_face_names = ["Khoa", "NhanPhan", "Minh"]

# Load known faces and known names
images_path = "./images/"
for name in known_face_names:
    image = face_recognition.load_image_file(f"{images_path}{name}.jpg")
    face_encoding = face_recognition.face_encodings(image)[0]
    known_face_encodings.append(face_encoding)

# Initialize webcam
video_capture = cv2.VideoCapture(0)

# Create or open an Excel workbook
excel_filename = 'attendance.xlsx'
if os.path.exists(excel_filename):
    workbook = load_workbook(excel_filename)
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Name"] + [f"Week {i+1}" for i in range(15)])
    for name in known_face_names:
        sheet.append([name] + [0]*15)
    workbook.save(excel_filename)

# Dictionary to store recognized names for each week
recognized_names_dict = {i: [] for i in range(1, 16)}
current_week = 1  # Initialize the current week

def face_distance_to_conf(face_distance, face_match_threshold=0.6):
    if face_distance > face_match_threshold:
        range = (1.0 - face_match_threshold)
        linear_val = (1.0 - face_distance) / (range * 2.0)
        return linear_val
    else:
        range = face_match_threshold
        linear_val = 1.0 - (face_distance / (range * 2.0))
        return linear_val + ((1.0 - linear_val) * math.pow((linear_val - 0.5) * 2, 0.2))

def gen_frames():
    global current_week
    while True:
        ret, frame = video_capture.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = np.ascontiguousarray(small_frame[:, :, ::-1])

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            name = "Unknown"
            confidence = face_distance_to_conf(face_distances[best_match_index], face_match_threshold=0.6)

            face_distance_threshold = 0.5
            if face_distances[best_match_index] < face_distance_threshold and confidence > 0.5:
                name = known_face_names[best_match_index]

            if name != "Unknown" and name not in recognized_names_dict[current_week]:
                recognized_names_dict[current_week].append(name)
                current_time = datetime.now().strftime("%H:%M:%S")
                with open('recognized_names.txt', 'a') as file:
                    file.write(f"{name},{current_time}\n")

                # Save to Excel
                workbook = load_workbook(excel_filename)
                sheet = workbook["Attendance"]
                for row in sheet.iter_rows(min_row=2, max_col=1):
                    if row[0].value == name:
                        sheet.cell(row=row[0].row, column=current_week+1).value = 1
                        break
                workbook.save(excel_filename)

                # Emit the updated attendance log to clients
                attendance_log = get_attendance_log()
                yield f"data: {json.dumps(attendance_log)}\n\n"

            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
            cv2.putText(frame, f"Confidence: {confidence:.2f}", (left, bottom + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 1)
            cv2.putText(frame, f"Threshold: {face_distance_threshold}", (left, bottom + 40), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 1)

        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    return render_template('index.html', students=known_face_names)

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/set_week', methods=['POST'])
def set_week():
    global current_week
    data = request.get_json()
    week = data.get('week', 1)
    current_week = int(week)
    # Reset the recognized names for the new week
    recognized_names_dict[current_week] = []
    print(f"Week set to: {current_week}")
    return jsonify({"status": "success", "current_week": current_week})

@app.route('/save_week', methods=['POST'])
def save_week():
    data = request.get_json()
    week = data.get('week', 1)
    recognized_names_dict[int(week)] = "saved"
    return jsonify({"status": "success"})

def get_attendance_log():
    workbook = load_workbook(excel_filename)
    sheet = workbook["Attendance"]
    attendance = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        weeks = row[1:]
        attendance[name] = weeks
    return attendance

@app.route('/attendance_log_stream')
def attendance_log_stream():
    @stream_with_context
    def generate():
        last_log = None
        while True:
            current_log = json.dumps(get_attendance_log())
            if current_log != last_log:
                yield f"data: {current_log}\n\n"
                last_log = current_log
            time.sleep(1)
    return Response(generate(), mimetype='text/event-stream')

@app.route('/reset_attendance', methods=['POST'])
def reset_attendance():
    workbook = load_workbook(excel_filename)
    sheet = workbook["Attendance"]
    for row in sheet.iter_rows(min_row=2, values_only=False):
        for cell in row[1:]:
            cell.value = 0
    workbook.save(excel_filename)
    for key in recognized_names_dict:
        recognized_names_dict[key] = []
    return jsonify({"status": "success"})

@app.route('/get_saved_weeks', methods=['GET'])
def get_saved_weeks():
    saved_weeks = [week for week, status in recognized_names_dict.items() if status == "saved"]
    return jsonify({"saved_weeks": saved_weeks})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
